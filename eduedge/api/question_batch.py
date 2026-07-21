from __future__ import annotations

import base64
import csv
import io
import re
from collections import Counter
from typing import Any

import frappe
from frappe import _
from frappe.utils import cint, cstr, flt

from eduedge.api.question_builder import DIFFICULTIES, EXAM_BODIES, QUESTION_TYPES
from eduedge.cbt.public_access import get_public_exam_capability_summary
from eduedge.eduedge.doctype.eduedge_cbt_question.eduedge_cbt_question import (
	BINARY_ANSWER_PRESETS,
	OBJECTIVE_TYPES,
	PLATFORM_BANK,
	SCHOOL_BANK,
	_require_question_author,
	option_label,
)
from eduedge.services.branch_context import get_allowed_school_branches, get_current_school_branch

QUESTION_DOCTYPE = "EduEdge CBT Question"
MAX_MANUAL_QUESTIONS = 50
MAX_UPLOAD_ROWS = 500
MAX_UPLOAD_BYTES = 5 * 1024 * 1024
ANSWER_COLUMNS = tuple(f"answer_{letter.lower()}" for letter in "ABCDEFGH")
QUESTION_TYPE_ALIASES = {
	"single": "Single Choice",
	"single choice": "Single Choice",
	"multiple": "Multiple Choice",
	"multiple choice": "Multiple Choice",
	"true false": "True/False",
	"true/false": "True/False",
	"yes no": "Yes/No",
	"yes/no": "Yes/No",
	"short": "Short Answer",
	"short answer": "Short Answer",
	"essay": "Essay",
	"numeric": "Numeric",
	"number": "Numeric",
}
HEADER_ALIASES = {
	"code": "question_code",
	"question_code": "question_code",
	"question_id": "question_code",
	"type": "question_type",
	"question_type": "question_type",
	"question": "question_text",
	"question_text": "question_text",
	"difficulty": "difficulty",
	"correct": "correct_answers",
	"correct_answer": "correct_answers",
	"correct_answers": "correct_answers",
	"answer_key": "answer_key",
	"marking_guide": "marking_guide",
	"mark": "default_mark",
	"default_mark": "default_mark",
	"negative_mark": "negative_mark",
	"notes": "notes",
}
for _letter in "ABCDEFGH":
	HEADER_ALIASES[f"answer_{_letter.lower()}"] = f"answer_{_letter.lower()}"
	HEADER_ALIASES[f"option_{_letter.lower()}"] = f"answer_{_letter.lower()}"


def _parse_payload(payload) -> dict:
	if isinstance(payload, str):
		return frappe.parse_json(payload) or {}
	return payload or {}


def _require_create_permission() -> None:
	_require_question_author()
	if not frappe.has_permission(QUESTION_DOCTYPE, "create"):
		frappe.throw(_("You are not permitted to create CBT questions."), frappe.PermissionError)


def _scope_options(public_access: dict) -> list[dict]:
	options = [{"value": SCHOOL_BANK, "label": _("School Question Bank")}]
	if public_access.get("capabilities", {}).get("author", {}).get("allowed"):
		options.append({"value": PLATFORM_BANK, "label": _("EduEdge Examination Bank")})
	return options


def _common_defaults() -> dict:
	current_branch = get_current_school_branch() or {}
	return {
		"ownership_scope": SCHOOL_BANK,
		"school_branch": current_branch.get("name") or "",
		"course": "",
		"course_label": "",
		"topic": "",
		"topic_label": "",
		"curriculum": "",
		"exam_body": "School Internal",
		"difficulty": "Moderate",
		"default_mark": 1,
		"negative_mark": 0,
	}


@frappe.whitelist()
def get_question_batch_context() -> dict:
	_require_create_permission()
	public_access = get_public_exam_capability_summary(frappe.session.user)
	current_branch = get_current_school_branch() or {}
	return {
		"user": {
			"name": frappe.session.user,
			"full_name": frappe.db.get_value("User", frappe.session.user, "full_name") or frappe.session.user,
		},
		"tenant_name": current_branch.get("company") or "",
		"current_branch": current_branch,
		"allowed_branches": get_allowed_school_branches(),
		"scope_options": _scope_options(public_access),
		"question_types": list(QUESTION_TYPES),
		"difficulties": list(DIFFICULTIES),
		"exam_bodies": list(EXAM_BODIES),
		"defaults": _common_defaults(),
		"limits": {
			"manual_questions": MAX_MANUAL_QUESTIONS,
			"upload_rows": MAX_UPLOAD_ROWS,
			"upload_bytes": MAX_UPLOAD_BYTES,
		},
	}


def _normalise_header(value: Any) -> str:
	key = re.sub(r"[^a-z0-9]+", "_", cstr(value).strip().lower()).strip("_")
	return HEADER_ALIASES.get(key, key)


def _cell_text(value: Any) -> str:
	if value is None:
		return ""
	if isinstance(value, bool):
		return "1" if value else "0"
	if isinstance(value, float) and value.is_integer():
		return str(int(value))
	return cstr(value).strip()


def _decode_upload(file_content: str) -> bytes:
	content = cstr(file_content or "").strip()
	if not content:
		frappe.throw(_("Select a CSV or XLSX file first."), frappe.ValidationError)
	if "," in content and content.lower().startswith("data:"):
		content = content.split(",", 1)[1]
	try:
		decoded = base64.b64decode(content, validate=True)
	except Exception:
		frappe.throw(_("The uploaded file content could not be decoded."), frappe.ValidationError)
	if len(decoded) > MAX_UPLOAD_BYTES:
		frappe.throw(_("The upload file cannot exceed 5 MB."), frappe.ValidationError)
	return decoded


def _parse_csv(content: bytes) -> list[dict]:
	try:
		text = content.decode("utf-8-sig")
	except UnicodeDecodeError:
		frappe.throw(_("CSV files must use UTF-8 encoding."), frappe.ValidationError)
	reader = csv.DictReader(io.StringIO(text))
	if not reader.fieldnames:
		frappe.throw(_("The CSV file has no header row."), frappe.ValidationError)
	raw_headers = list(reader.fieldnames)
	headers = [_normalise_header(header) for header in raw_headers]
	rows = []
	for row_number, values in enumerate(reader, start=2):
		row = {
			header: _cell_text(values.get(raw_header))
			for raw_header, header in zip(raw_headers, headers)
			if header
		}
		if any(value for value in row.values()):
			row["_row_number"] = row_number
			rows.append(row)
	return rows


def _parse_xlsx(content: bytes) -> list[dict]:
	try:
		from openpyxl import load_workbook
	except ImportError:
		frappe.throw(
			_("XLSX support is unavailable on this bench. Upload the same template as CSV."),
			frappe.ValidationError,
		)
	try:
		workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
	except Exception:
		frappe.throw(_("The XLSX workbook could not be read."), frappe.ValidationError)
	try:
		worksheet = workbook.active
		iterator = worksheet.iter_rows(values_only=True)
		try:
			raw_headers = next(iterator)
		except StopIteration:
			frappe.throw(_("The XLSX workbook is empty."), frappe.ValidationError)
		headers = [_normalise_header(value) for value in raw_headers]
		if not any(headers):
			frappe.throw(_("The XLSX workbook has no header row."), frappe.ValidationError)
		rows = []
		for row_number, values in enumerate(iterator, start=2):
			row = {
				headers[index]: _cell_text(values[index]) if index < len(values) else ""
				for index in range(len(headers))
				if headers[index]
			}
			if any(value for value in row.values()):
				row["_row_number"] = row_number
				rows.append(row)
		return rows
	finally:
		workbook.close()


def _parse_upload(file_name: str, content: bytes) -> list[dict]:
	extension = cstr(file_name or "").lower().rsplit(".", 1)[-1]
	if extension == "csv":
		rows = _parse_csv(content)
	elif extension == "xlsx":
		rows = _parse_xlsx(content)
	else:
		frappe.throw(_("Only CSV and XLSX question files are supported."), frappe.ValidationError)
	if not rows:
		frappe.throw(_("The upload file contains no question rows."), frappe.ValidationError)
	if len(rows) > MAX_UPLOAD_ROWS:
		frappe.throw(
			_("The upload contains {0} rows. The maximum is {1} rows per import.").format(
				len(rows), MAX_UPLOAD_ROWS
			),
			frappe.ValidationError,
		)
	return rows


def _normalise_question_type(value: Any) -> str:
	text = cstr(value or "Single Choice").strip()
	question_type = QUESTION_TYPE_ALIASES.get(text.lower(), text)
	if question_type not in QUESTION_TYPES:
		frappe.throw(_("Select a valid Question Type."), frappe.ValidationError)
	return question_type


def _split_correct_answers(value: Any) -> list[str]:
	text = cstr(value or "").strip()
	if not text:
		return []
	return [token.strip() for token in re.split(r"[,|;/]+", text) if token.strip()]


def _upload_answer_options(row: dict, question_type: str) -> list[dict]:
	if question_type in BINARY_ANSWER_PRESETS:
		answers = list(BINARY_ANSWER_PRESETS[question_type])
		provided = [_cell_text(row.get(column)) for column in ANSWER_COLUMNS]
		provided = [value for value in provided if value]
		if provided and [value.lower() for value in provided] != [value.lower() for value in answers]:
			frappe.throw(
				_("{0} uses the fixed answers {1}.").format(question_type, " and ".join(answers)),
				frappe.ValidationError,
			)
	else:
		answers = []
		blank_seen = False
		for column in ANSWER_COLUMNS:
			answer = _cell_text(row.get(column))
			if not answer:
				blank_seen = True
				continue
			if blank_seen:
				frappe.throw(
					_("Answer columns must start at Answer A and remain continuous without gaps."),
					frappe.ValidationError,
				)
			answers.append(answer)

	if question_type not in OBJECTIVE_TYPES:
		if answers or _split_correct_answers(row.get("correct_answers")):
			frappe.throw(
				_("Answer choice columns are only valid for objective question types."),
				frappe.ValidationError,
			)
		return []
	if len(answers) < 2:
		frappe.throw(_("Objective questions require at least two Answers."), frappe.ValidationError)

	correct_tokens = _split_correct_answers(row.get("correct_answers"))
	correct_indexes: set[int] = set()
	for token in correct_tokens:
		normalised = token.strip().lower()
		matched = None
		for index, answer in enumerate(answers):
			if normalised in {option_label(index + 1).lower(), answer.strip().lower()}:
				matched = index
				break
		if matched is None:
			frappe.throw(
				_("Correct Answer {0} does not match an option label or answer text.").format(
					frappe.bold(token)
				),
				frappe.ValidationError,
			)
		correct_indexes.add(matched)

	return [
		{
			"option_text": answer,
			"is_correct": 1 if index in correct_indexes else 0,
		}
		for index, answer in enumerate(answers)
	]


def _manual_answer_options(row: dict) -> list[dict]:
	return [
		{
			"option_text": _cell_text(answer.get("option_text")),
			"is_correct": cint(answer.get("is_correct")),
		}
		for answer in (row.get("options") or [])
	]


def _build_question_doc(common: dict, row: dict, *, upload: bool = False):
	question_type = _normalise_question_type(row.get("question_type"))
	doc = frappe.new_doc(QUESTION_DOCTYPE)
	doc.question_code = cstr(row.get("question_code") or "").strip().upper()
	doc.ownership_scope = common.get("ownership_scope") or SCHOOL_BANK
	doc.school_branch = common.get("school_branch") or None
	doc.version_number = 1
	doc.course = common.get("course") or None
	doc.topic = common.get("topic") or None
	doc.curriculum = cstr(common.get("curriculum") or "").strip()
	doc.exam_body = common.get("exam_body") or "School Internal"
	doc.difficulty = row.get("difficulty") or common.get("difficulty") or "Moderate"
	doc.question_type = question_type
	doc.question_text = row.get("question_text") or row.get("question") or ""
	doc.answer_key = row.get("answer_key") or ""
	doc.marking_guide = row.get("marking_guide") or ""
	doc.default_mark = flt(row.get("default_mark")) or flt(common.get("default_mark")) or 1
	doc.negative_mark = (
		flt(row.get("negative_mark"))
		if cstr(row.get("negative_mark")).strip() != ""
		else flt(common.get("negative_mark"))
	)
	doc.notes = row.get("notes") or ""
	doc.status = "Draft"

	options = _upload_answer_options(row, question_type) if upload else _manual_answer_options(row)
	for answer in options:
		doc.append("options", answer)
	return doc


def _normalise_common(payload: dict) -> dict:
	common = {**_common_defaults(), **(payload or {})}
	common["ownership_scope"] = common.get("ownership_scope") or SCHOOL_BANK
	if common["ownership_scope"] == PLATFORM_BANK:
		common["school_branch"] = None
	return common


def _duplicate_codes(rows: list[dict]) -> tuple[set[str], set[str]]:
	codes = [cstr(row.get("question_code") or "").strip().upper() for row in rows]
	within_batch = {code for code, count in Counter(code for code in codes if code).items() if count > 1}
	unique_codes = sorted({code for code in codes if code})
	existing = set()
	if unique_codes:
		existing = set(
			frappe.get_all(
				QUESTION_DOCTYPE,
				filters={"name": ["in", unique_codes]},
				pluck="name",
				limit_page_length=0,
			)
		)
	return within_batch, existing


@frappe.whitelist()
def save_question_batch(common, questions, source: str | None = None) -> dict:
	_require_create_permission()
	common_values = _normalise_common(_parse_payload(common))
	question_rows = _parse_payload(questions)
	if not isinstance(question_rows, list) or not question_rows:
		frappe.throw(_("Add at least one question before saving."), frappe.ValidationError)
	limit = MAX_UPLOAD_ROWS if source == "upload" else MAX_MANUAL_QUESTIONS
	if len(question_rows) > limit:
		frappe.throw(
			_("This operation supports a maximum of {0} questions.").format(limit),
			frappe.ValidationError,
		)

	within_batch, existing = _duplicate_codes(question_rows)
	if within_batch:
		frappe.throw(
			_("Duplicate Question Codes in this batch: {0}").format(", ".join(sorted(within_batch))),
			frappe.ValidationError,
		)
	if existing:
		frappe.throw(
			_("These Question Codes already exist: {0}").format(", ".join(sorted(existing))),
			frappe.ValidationError,
		)

	docs = []
	for index, row in enumerate(question_rows, start=1):
		try:
			doc = _build_question_doc(common_values, row, upload=source == "upload")
			doc.run_method("validate")
		except frappe.PermissionError:
			raise
		except Exception as exc:
			frappe.throw(
				_("Question {0}: {1}").format(index, cstr(exc)),
				frappe.ValidationError,
			)
		docs.append(doc)

	created = []
	for doc in docs:
		doc.insert()
		created.append(doc.name)
	return {
		"created": created,
		"count": len(created),
		"status": "Draft",
	}


@frappe.whitelist()
def preview_question_upload(file_name: str, file_content: str, common) -> dict:
	_require_create_permission()
	common_values = _normalise_common(_parse_payload(common))
	raw_rows = _parse_upload(file_name, _decode_upload(file_content))
	within_batch, existing = _duplicate_codes(raw_rows)
	preview = []

	for position, row in enumerate(raw_rows, start=1):
		row_number = cint(row.get("_row_number")) or position + 1
		code = cstr(row.get("question_code") or "").strip().upper()
		error = ""
		normalised_type = row.get("question_type")
		if code and code in within_batch:
			error = _("Question Code is repeated in this file.")
		elif code and code in existing:
			error = _("Question Code already exists.")
		else:
			try:
				doc = _build_question_doc(common_values, row, upload=True)
				doc.run_method("validate")
				normalised_type = doc.question_type
			except frappe.PermissionError:
				raise
			except Exception as exc:
				error = cstr(exc)

		preview.append(
			{
				"row_number": row_number,
				"question_code": code,
				"question_type": normalised_type,
				"question_text": cstr(row.get("question_text") or row.get("question") or "")[:180],
				"valid": not bool(error),
				"error": error,
			}
		)

	return {
		"file_name": file_name,
		"rows": preview,
		"total_rows": len(preview),
		"valid_rows": sum(1 for row in preview if row["valid"]),
		"error_rows": sum(1 for row in preview if not row["valid"]),
		"can_import": bool(preview) and all(row["valid"] for row in preview),
	}
